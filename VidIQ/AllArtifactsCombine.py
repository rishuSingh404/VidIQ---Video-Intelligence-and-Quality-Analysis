# -*- coding: utf-8 -*-
"""
Created on Wed Jan  8 03:54:41 2020

@author: user
"""

import openpyxl

Name = ['BigBuckBunny', 'BirdsInCage', 'BQTerrace', 'CrowdRun', 'DanceKiss', 'ElFuente1',
        'ElFuente2', 'FoxBird', 'Kimono1', 'OldTownCross', 'Seeking', 'Tennis']
FrameNo = [150, 180, 180, 150, 150, 180, 180, 150, 132, 150, 150, 120]
Format = ["_H264_","_SDH264HD_"]
File = ['BBB', 'BIC', 'BQT', 'CRN', 'DKS', 'EF1', 'EF2', 'FOX', 'KIM', 'OTC', 'SEE', 'TEN']

for vid in range(0,12):
    for f in range(0,2):
        for level in range(1,5):
            if f == 0:
                filename1 = 'Frame/' + File[vid] + '/' + File[vid] + '_' + str(level) + '.xlsx'
            elif f == 1:
                filename1 = 'Frame/' + File[vid] + '/' + File[vid] + '_N' + str(level) + '.xlsx'
            filename2 = 'Jerkiness/' + str(Name[vid]) + str(Format[f]) + str(level) + '_Jerkiness.xlsx'
            filename3 = 'Flickering/' + str(Name[vid]) + str(Format[f]) + str(level) + '_Flickering.xlsx'
            filename4 = 'MosquitoNoise/' + str(Name[vid]) + str(Format[f]) + str(level) + '_MosquitoNoise.xlsx'
            
            book = openpyxl.Workbook()
            sheet = book.active
            book1 = openpyxl.load_workbook(filename1)
            sheet1 = book1.active
            book2 = openpyxl.load_workbook(filename2)
            sheet2 = book2.active
            book3 = openpyxl.load_workbook(filename3)
            sheet3 = book3.active
            book4 = openpyxl.load_workbook(filename4)
            sheet4 = book4.active
            
            sheet['A1'] = 'Frame'
            sheet['B1'] = 'G-Noise'
            sheet['C1'] = 'Blurry'
            sheet['D1'] = 'Sharperness'
            sheet['E1'] = 'Blocky'
            sheet['F1'] = 'Freeze-100%'
            sheet['G1'] = 'Freeze-90%'
            sheet['H1'] = 'Freeze-85%'
            sheet['I1'] = 'Freeze-80%'
            sheet['J1'] = 'Freeze-75%'
            sheet['K1'] = 'Freeze-Percent'
            sheet['L1'] = 'Jerkiness'
            sheet['M1'] = 'Flickering'
            sheet['N1'] = 'MosquitoNoise'
            for i in range(1, 181):
                sheet['A' + str(i + 1)] = i
            
            for i in range(2, FrameNo[vid]+2):
                sheet['B' + str(i)] = sheet1['B' + str(i)].value
                sheet['C' + str(i)] = sheet1['C' + str(i)].value
                sheet['D' + str(i)] = sheet1['D' + str(i)].value
                sheet['E' + str(i)] = sheet1['E' + str(i)].value
                sheet['F' + str(i)] = sheet1['F' + str(i)].value
                sheet['G' + str(i)] = sheet1['G' + str(i)].value
                sheet['H' + str(i)] = sheet1['H' + str(i)].value
                sheet['I' + str(i)] = sheet1['I' + str(i)].value
                sheet['J' + str(i)] = sheet1['J' + str(i)].value
                sheet['K' + str(i)] = sheet1['K' + str(i)].value
                sheet['L' + str(i)] = sheet2['D' + str(i)].value
                sheet['M' + str(i)] = sheet3['B' + str(i)].value
                sheet['N' + str(i)] = sheet4['D' + str(i)].value
            
            if FrameNo[vid] < 180:
                for i in range(FrameNo[vid] + 2, 182):
                    sheet['B' + str(i)] = 0
                    sheet['C' + str(i)] = 0
                    sheet['D' + str(i)] = 0
                    sheet['E' + str(i)] = 0
                    sheet['F' + str(i)] = 0
                    sheet['G' + str(i)] = 0
                    sheet['H' + str(i)] = 0
                    sheet['I' + str(i)] = 0
                    sheet['J' + str(i)] = 0
                    sheet['K' + str(i)] = 0
                    sheet['L' + str(i)] = 0
                    sheet['M' + str(i)] = 0
                    sheet['N' + str(i)] = 0
            
            book.save('All Artifacts/' + str(Name[vid]) + str(Format[f]) + str(level) + '_AllArtifacts.xlsx')
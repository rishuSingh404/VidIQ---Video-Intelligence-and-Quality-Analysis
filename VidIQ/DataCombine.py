# -*- coding: utf-8 -*-
"""
Created on Tue Jan 14 16:03:53 2020

@author: user
"""

import openpyxl

TrainName = ['BigBuckBunny_H264_1', 'BigBuckBunny_H264_2', 'BigBuckBunny_H264_3', 'BigBuckBunny_SDH264HD_1', 'BigBuckBunny_SDH264HD_2', 'BigBuckBunny_SDH264HD_3',
             'BirdsInCage_H264_1', 'BirdsInCage_H264_4', 'BirdsInCage_SDH264HD_1', 'BirdsInCage_SDH264HD_2', 'BirdsInCage_SDH264HD_3', 'BirdsInCage_SDH264HD_4',
             'BQTerrace_H264_3', 'BQTerrace_H264_4', 'BQTerrace_SDH264HD_1', 'BQTerrace_SDH264HD_2', 'BQTerrace_SDH264HD_3', 'BQTerrace_SDH264HD_4',
             'CrowdRun_H264_2', 'CrowdRun_H264_3', 'CrowdRun_H264_4', 'CrowdRun_SDH264HD_1', 'CrowdRun_SDH264HD_2', 'CrowdRun_SDH264HD_3', 'CrowdRun_SDH264HD_4',
             'DanceKiss_H264_1', 'DanceKiss_H264_2', 'DanceKiss_H264_3', 'DanceKiss_H264_4', 'DanceKiss_SDH264HD_1', 'DanceKiss_SDH264HD_2', 'DanceKiss_SDH264HD_3',
             'ElFuente1_H264_1', 'ElFuente1_H264_2', 'ElFuente1_H264_3', 'ElFuente1_SDH264HD_1', 'ElFuente1_SDH264HD_3', 'ElFuente1_SDH264HD_4',
             'ElFuente2_H264_1', 'ElFuente2_H264_3', 'ElFuente2_H264_4', 'ElFuente2_SDH264HD_2', 'ElFuente2_SDH264HD_3', 'ElFuente2_SDH264HD_4',
             'FoxBird_H264_2', 'FoxBird_H264_4', 'FoxBird_SDH264HD_1', 'FoxBird_SDH264HD_2', 'FoxBird_SDH264HD_3', 'FoxBird_SDH264HD_4',
             'Kimono1_H264_1', 'Kimono1_H264_2', 'Kimono1_H264_3', 'Kimono1_H264_4', 'Kimono1_SDH264HD_1', 'Kimono1_SDH264HD_3',
             'OldTownCross_H264_1', 'OldTownCross_H264_2', 'OldTownCross_H264_4', 'OldTownCross_SDH264HD_1', 'OldTownCross_SDH264HD_2', 'OldTownCross_SDH264HD_3', 'OldTownCross_SDH264HD_4',
             'Seeking_H264_1', 'Seeking_H264_2', 'Seeking_H264_3', 'Seeking_H264_4', 'Seeking_SDH264HD_2', 'Seeking_SDH264HD_4',
             'Tennis_H264_1', 'Tennis_H264_2', 'Tennis_H264_3', 'Tennis_H264_4', 'Tennis_SDH264HD_2', 'Tennis_SDH264HD_3', 'Tennis_SDH264HD_4']

TestName = ['BigBuckBunny_H264_4', 'BigBuckBunny_SDH264HD_4', 'BirdsInCage_H264_2', 'BirdsInCage_H264_3',
             'BQTerrace_H264_1', 'BQTerrace_H264_2', 'CrowdRun_H264_1', 'DanceKiss_SDH264HD_4',
             'ElFuente1_H264_4', 'ElFuente1_SDH264HD_2', 'ElFuente2_H264_2', 'ElFuente2_SDH264HD_1',
             'FoxBird_H264_1', 'FoxBird_H264_3', 'Kimono1_SDH264HD_2', 'Kimono1_SDH264HD_4',
             'OldTownCross_H264_3', 'Seeking_SDH264HD_1', 'Seeking_SDH264HD_3', 'Tennis_SDH264HD_1']

bookTrain = openpyxl.Workbook()
sheetTrain = bookTrain.active
sheetTrain['A1'] = 'Video'
sheetTrain['B1'] = 'G-Noise'
sheetTrain['C1'] = 'Blurry'
sheetTrain['D1'] = 'Sharperness'
sheetTrain['E1'] = 'Blocky'
sheetTrain['F1'] = 'Freeze-100%'
sheetTrain['G1'] = 'Freeze-90%'
sheetTrain['H1'] = 'Freeze-85%'
sheetTrain['I1'] = 'Freeze-80%'
sheetTrain['J1'] = 'Freeze-75%'
sheetTrain['K1'] = 'Freeze-Percent'
sheetTrain['L1'] = 'Jerkiness'
sheetTrain['M1'] = 'Flickering'
sheetTrain['N1'] = 'MosquitoNoise'
bookTest = openpyxl.Workbook()
sheetTest = bookTest.active
sheetTest['A1'] = 'Video'
sheetTest['B1'] = 'G-Noise'
sheetTest['C1'] = 'Blurry'
sheetTest['D1'] = 'Sharperness'
sheetTest['E1'] = 'Blocky'
sheetTest['F1'] = 'Freeze-100%'
sheetTest['G1'] = 'Freeze-90%'
sheetTest['H1'] = 'Freeze-85%'
sheetTest['I1'] = 'Freeze-80%'
sheetTest['J1'] = 'Freeze-75%'
sheetTest['K1'] = 'Freeze-Percent'
sheetTest['L1'] = 'Jerkiness'
sheetTest['M1'] = 'Flickering'
sheetTest['N1'] = 'MosquitoNoise'

for i in range(0, 76):
    filename = 'All Artifacts/' + TrainName[i] +  '_AllArtifacts.xlsx'   
    book1 = openpyxl.load_workbook(filename)
    sheet1 = book1.active
    for j in range((i * 180 + 2), ((i + 1) * 180 + 2)):
        sheetTrain['A' + str(j)] = TrainName[i]   
        sheetTrain['B' + str(j)] = sheet1['B' + str(j - 180 * i)].value
        sheetTrain['C' + str(j)] = sheet1['C' + str(j - 180 * i)].value
        sheetTrain['D' + str(j)] = sheet1['D' + str(j - 180 * i)].value   
        sheetTrain['E' + str(j)] = sheet1['E' + str(j - 180 * i)].value   
        sheetTrain['F' + str(j)] = sheet1['F' + str(j - 180 * i)].value
        sheetTrain['G' + str(j)] = sheet1['G' + str(j - 180 * i)].value   
        sheetTrain['H' + str(j)] = sheet1['H' + str(j - 180 * i)].value
        sheetTrain['I' + str(j)] = sheet1['I' + str(j - 180 * i)].value   
        sheetTrain['J' + str(j)] = sheet1['J' + str(j - 180 * i)].value
        sheetTrain['K' + str(j)] = sheet1['K' + str(j - 180 * i)].value   
        sheetTrain['L' + str(j)] = sheet1['L' + str(j - 180 * i)].value
        sheetTrain['M' + str(j)] = sheet1['M' + str(j - 180 * i)].value   
        sheetTrain['N' + str(j)] = sheet1['N' + str(j - 180 * i)].value
        
bookTrain.save('TrainingData.xlsx')

for i in range(0, 20):      
    filename = 'All Artifacts/' + TestName[i] +  '_AllArtifacts.xlsx'
    book1 = openpyxl.load_workbook(filename)
    sheet1 = book1.active
    for j in range((i * 180 + 2), ((i + 1) * 180 + 2)):
        sheetTest['A' + str(j)] = TestName[i]   
        sheetTest['B' + str(j)] = sheet1['B' + str(j - 180 * i)].value
        sheetTest['C' + str(j)] = sheet1['C' + str(j - 180 * i)].value
        sheetTest['D' + str(j)] = sheet1['D' + str(j - 180 * i)].value   
        sheetTest['E' + str(j)] = sheet1['E' + str(j - 180 * i)].value   
        sheetTest['F' + str(j)] = sheet1['F' + str(j - 180 * i)].value
        sheetTest['G' + str(j)] = sheet1['G' + str(j - 180 * i)].value   
        sheetTest['H' + str(j)] = sheet1['H' + str(j - 180 * i)].value
        sheetTest['I' + str(j)] = sheet1['I' + str(j - 180 * i)].value   
        sheetTest['J' + str(j)] = sheet1['J' + str(j - 180 * i)].value
        sheetTest['K' + str(j)] = sheet1['K' + str(j - 180 * i)].value   
        sheetTest['L' + str(j)] = sheet1['L' + str(j - 180 * i)].value
        sheetTest['M' + str(j)] = sheet1['M' + str(j - 180 * i)].value   
        sheetTest['N' + str(j)] = sheet1['N' + str(j - 180 * i)].value

bookTest.save('TestingData.xlsx')